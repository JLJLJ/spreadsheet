"""Excel文件处理模块"""
import os
from pathlib import Path
from typing import Any, List, Dict, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import json

from database import SHEETS_DIR


def create_empty_sheet(file_name: str) -> str:
    """创建空白Excel文件"""
    file_path = str(SHEETS_DIR / f"{file_name}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # 不需要预填充单元格，openpyxl会自动处理
    wb.save(file_path)
    wb.close()
    return file_path


def load_sheet_data(file_path: str) -> Dict[str, Any]:
    """加载Excel文件数据，转换为Univer可用的格式"""
    wb = load_workbook(file_path)
    ws = wb.active

    # 安全获取sheet标题，处理可能的编码问题
    try:
        sheet_title = ws.title if ws.title else "Sheet1"
        # 确保是有效的UTF-8字符串
        if isinstance(sheet_title, bytes):
            sheet_title = sheet_title.decode('utf-8', errors='replace')
    except Exception:
        sheet_title = "Sheet1"

    # 获取数据范围
    max_row = max(ws.max_row, 100)
    max_col = max(ws.max_column, 26)

    # 读取单元格数据
    cell_data = {}
    styles = {}
    merges = []

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell_key = f"{row-1}_{col-1}"  # 转为0索引

            # 获取样式（先获取样式，因为空单元格也可能有边框等样式）
            style = extract_cell_style(cell)

            # 获取值
            value = cell.value
            # 修复：即使单元格为空，只要有样式（如边框），也要创建单元格条目
            if value is not None or (style and len(style) > 0):
                cell_data[cell_key] = {}

                if value is not None:
                    cell_data[cell_key]["v"] = value
                    cell_data[cell_key]["t"] = get_cell_type(value)  # type: s=string, n=number, b=boolean

                # 添加样式（包括边框）
                if style:
                    cell_data[cell_key].update(style)

    # 获取合并单元格信息
    for merge_range in ws.merged_cells.ranges:
        merges.append({
            "startRow": merge_range.min_row - 1,
            "endRow": merge_range.max_row - 1,
            "startColumn": merge_range.min_col - 1,
            "endColumn": merge_range.max_col - 1,
        })

    # 获取列宽
    col_widths = {}
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        if col_letter in ws.column_dimensions:
            width = ws.column_dimensions[col_letter].width
            if width:
                col_widths[col - 1] = int(width * 7)  # 转换为像素

    # 获取行高
    row_heights = {}
    for row in range(1, max_row + 1):
        if row in ws.row_dimensions:
            height = ws.row_dimensions[row].height
            if height:
                row_heights[row - 1] = int(height)

    return {
        "id": Path(file_path).stem,
        "name": sheet_title,
        "cellData": cell_data,
        "mergeData": merges,
        "columnData": col_widths,
        "rowData": row_heights,
        "rowCount": max_row,
        "columnCount": max_col,
    }


def get_cell_type(value: Any) -> str:
    """获取单元格值类型"""
    if isinstance(value, bool):
        return "b"
    elif isinstance(value, (int, float)):
        return "n"
    else:
        return "s"


def extract_cell_style(cell) -> Optional[Dict]:
    """提取单元格样式（完整版：字体、颜色、对齐、边框等）"""
    style = {}

    # 字体
    if cell.font:
        font = cell.font
        if font.bold:
            style["bl"] = 1
        if font.italic:
            style["it"] = 1
        if font.underline:
            style["ul"] = {"s": 1}  # single underline
        if font.strike:
            style["st"] = 1
        # 字体颜色
        try:
            if font.color and font.color.rgb and isinstance(font.color.rgb, str):
                rgb_val = font.color.rgb
                if len(rgb_val) >= 6 and all(c in '0123456789ABCDEFabcdef' for c in rgb_val):
                    style["cl"] = {"rgb": rgb_val}
        except Exception:
            pass
        if font.size:
            style["fs"] = font.size
        if font.name:
            style["ff"] = font.name

    # 背景色
    try:
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            rgb = cell.fill.fgColor.rgb
            if isinstance(rgb, str) and len(rgb) >= 6 and rgb != "00000000":
                if all(c in '0123456789ABCDEFabcdef' for c in rgb):
                    style["bg"] = {"rgb": rgb}
    except Exception:
        pass

    # 对齐方式
    if cell.alignment:
        alignment = cell.alignment
        # 水平对齐: left, center, right
        if alignment.horizontal:
            h_map = {"left": "left", "center": "center", "right": "right"}
            if alignment.horizontal in h_map:
                style["ht"] = h_map[alignment.horizontal]
        # 垂直对齐: top, middle, bottom
        if alignment.vertical:
            v_map = {"top": "top", "center": "middle", "bottom": "bottom"}
            if alignment.vertical in v_map:
                style["vt"] = v_map[alignment.vertical]
        # 文字换行
        if alignment.wrap_text:
            style["tb"] = "2"  # wrap text

    # 边框
    if cell.border:
        border = cell.border
        border_style = {}
        for side_name, side_key in [("top", "t"), ("bottom", "b"), ("left", "l"), ("right", "r")]:
            side = getattr(border, side_name, None)
            if side and side.style:
                # 安全获取边框颜色
                border_color = "000000"  # 默认黑色
                try:
                    if side.color and side.color.rgb:
                        color_value = side.color.rgb
                        # 确保是字符串并且是有效的十六进制颜色
                        if isinstance(color_value, str):
                            # 去除非十六进制字符
                            cleaned = ''.join(c for c in color_value if c in '0123456789ABCDEFabcdef')
                            if len(cleaned) >= 6:
                                border_color = cleaned[:8]  # 最多取8位（ARGB）
                        elif isinstance(color_value, int):
                            # 如果是整数，转换为十六进制
                            border_color = format(color_value, '06X')
                except Exception:
                    pass

                border_style[side_key] = {
                    "s": 1 if side.style == "thin" else 2,  # 1=thin, 2=thick
                    "cl": {"rgb": border_color}
                }
        if border_style:
            style["bd"] = border_style

    return style if style else None


def update_cell(file_path: str, row: int, col: int, value: Any, style: Optional[Dict] = None):
    """更新单个单元格并保存"""
    wb = load_workbook(file_path)
    ws = wb.active

    cell = ws.cell(row=row + 1, column=col + 1)  # openpyxl使用1索引
    cell.value = value

    # 应用样式
    if style:
        apply_cell_style(cell, style)

    wb.save(file_path)


def apply_cell_style(cell, style: Dict):
    """应用样式到单元格（完整版：字体、颜色、对齐、边框等）"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # 字体样式
    font_kwargs = {}
    if "bl" in style and style["bl"]:
        font_kwargs["bold"] = True
    if "it" in style and style["it"]:
        font_kwargs["italic"] = True
    if "ul" in style:
        font_kwargs["underline"] = "single"
    if "st" in style and style["st"]:
        font_kwargs["strike"] = True
    if "fs" in style:
        font_kwargs["size"] = style["fs"]
    if "ff" in style:
        font_kwargs["name"] = style["ff"]
    if "cl" in style and "rgb" in style["cl"]:
        font_kwargs["color"] = style["cl"]["rgb"]
    if font_kwargs:
        cell.font = Font(**font_kwargs)

    # 背景色
    if "bg" in style and "rgb" in style["bg"]:
        rgb = style["bg"]["rgb"]
        cell.fill = PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")

    # 对齐方式
    alignment_kwargs = {}
    if "ht" in style:  # 水平对齐
        alignment_kwargs["horizontal"] = style["ht"]
    if "vt" in style:  # 垂直对齐
        alignment_kwargs["vertical"] = style["vt"]
    if "tb" in style and style["tb"] == "2":  # 文字换行
        alignment_kwargs["wrap_text"] = True
    if alignment_kwargs:
        cell.alignment = Alignment(**alignment_kwargs)

    # 边框
    if "bd" in style:
        bd = style["bd"]
        sides = {}
        for side_key, side_name in [("t", "top"), ("b", "bottom"), ("l", "left"), ("r", "right")]:
            if side_key in bd:
                border_info = bd[side_key]
                border_style = "thin" if border_info.get("s") == 1 else "thick"
                border_color = border_info.get("cl", {}).get("rgb", "000000")
                sides[side_name] = Side(style=border_style, color=border_color)
        if sides:
            cell.border = Border(**sides)


def batch_update_cells(file_path: str, updates: List[Dict]):
    """批量更新单元格"""
    wb = load_workbook(file_path)
    ws = wb.active

    for update in updates:
        row = update["row"] + 1
        col = update["col"] + 1
        value = update.get("value")
        style = update.get("style")

        cell = ws.cell(row=row, column=col)
        cell.value = value

        if style:
            apply_cell_style(cell, style)

    wb.save(file_path)


def save_sheet_from_univer(file_path: str, sheet_data: Dict):
    """从Univer格式保存为Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_data.get("name", "Sheet1")

    cell_data = sheet_data.get("cellData", {})

    for cell_key, cell_info in cell_data.items():
        parts = cell_key.split("_")
        if len(parts) == 2:
            row = int(parts[0]) + 1
            col = int(parts[1]) + 1
            cell = ws.cell(row=row, column=col)
            cell.value = cell_info.get("v", "")

            if "s" in cell_info:
                apply_cell_style(cell, cell_info["s"])

    # 应用合并单元格
    merges = sheet_data.get("mergeData", [])
    for merge in merges:
        start_row = merge["startRow"] + 1
        end_row = merge["endRow"] + 1
        start_col = merge["startColumn"] + 1
        end_col = merge["endColumn"] + 1
        ws.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=end_row,
            end_column=end_col
        )

    wb.save(file_path)


def import_excel(source_path: str, target_name: str) -> str:
    """导入外部Excel文件"""
    target_path = SHEETS_DIR / f"{target_name}.xlsx"

    # 直接复制文件
    import shutil
    shutil.copy(source_path, target_path)

    return str(target_path)


def update_column_width(file_path: str, col: int, width: int):
    """更新列宽"""
    wb = load_workbook(file_path)
    ws = wb.active

    col_letter = get_column_letter(col + 1)  # 转为1索引
    ws.column_dimensions[col_letter].width = width / 7  # 像素转为Excel单位

    wb.save(file_path)
    wb.close()


def update_row_height(file_path: str, row: int, height: int):
    """更新行高"""
    wb = load_workbook(file_path)
    ws = wb.active

    ws.row_dimensions[row + 1].height = height  # 转为1索引

    wb.save(file_path)
    wb.close()


def batch_update_dimensions(file_path: str, col_widths: Dict = None, row_heights: Dict = None):
    """批量更新列宽和行高"""
    wb = load_workbook(file_path)
    ws = wb.active

    if col_widths:
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(int(col_idx) + 1)
            ws.column_dimensions[col_letter].width = width / 7  # 像素转为Excel单位

    if row_heights:
        for row_idx, height in row_heights.items():
            ws.row_dimensions[int(row_idx) + 1].height = height

    wb.save(file_path)
    wb.close()
